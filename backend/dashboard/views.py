# backend/dashboard/views.py - COMPLETE FIXED VERSION
from xlsxwriter.utility import xl_rowcol_to_cell
from django.conf import settings
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.authtoken.models import Token
from django.utils import timezone
from rest_framework import viewsets, status, generics, mixins
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, JsonResponse, HttpRequest
from django.contrib.auth.models import User
from django.db.models import Q, Count, Sum, Avg
from django.db import IntegrityError
import pandas as pd
import numpy as np
from io import BytesIO, StringIO
import json
import os
import uuid
from datetime import datetime, timedelta
import csv
import xlsxwriter
import openpyxl
from openpyxl.utils import get_column_letter
import traceback
from openpyxl import load_workbook

from .models import (
    OutcomeDescription, CallDataFile, ProcessedData,
    GeneratedReport, ReportTemplate, Campaign
)
from .serializers import (
    OutcomeDescriptionSerializer, CallDataFileSerializer,
    ProcessedDataSerializer, GeneratedReportSerializer,
    ReportTemplateSerializer, FileUploadSerializer, CampaignSerializer
)


# ===========================================================
# AUTH VIEWS
# ===========================================================

class CustomAuthToken(ObtainAuthToken):
    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(
            data=request.data, context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)
        return Response({
            'token': token.key,
            'user_id': user.pk,
            'username': user.username,
            'email': user.email
        })


@api_view(['POST'])
@permission_classes([AllowAny])
def register_user(request):
    """Register a new user"""
    username = request.data.get('username')
    email = request.data.get('email')
    password = request.data.get('password')

    if not all([username, email, password]):
        return Response({'error': 'All fields are required'}, status=400)
    if User.objects.filter(username=username).exists():
        return Response({'error': 'Username already exists'}, status=400)
    if User.objects.filter(email=email).exists():
        return Response({'error': 'Email already exists'}, status=400)

    user = User.objects.create_user(username=username, email=email, password=password)
    user.is_active = True
    user.save()
    token = Token.objects.create(user=user)

    return Response({
        'message': 'User created successfully',
        'token': token.key,
        'user_id': user.id,
        'username': user.username,
        'email': user.email
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def verify_token(request):
    """Verify if the token is valid"""
    return Response({
        'valid': True,
        'user': {
            'username': request.user.username,
            'email': request.user.email,
        }
    })


# ===========================================================
# DATA PROCESSOR
# ===========================================================

class SimpleDataProcessor:
    @staticmethod
    def process_call_data(file_path, user, file_type='excel', delimiter=',', has_headers=True):
        """Process call data file and add Description column after last_outcome"""
        try:
            print(f"📁 Processing file: {file_path}")
            file_ext = os.path.splitext(file_path)[1].lower()

            if file_ext == '.csv':
                print("🔍 Examining CSV file structure...")
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    first_line = f.readline().strip()
                    second_line = f.readline().strip()
                    header_cols = first_line.split(delimiter)
                    print(f"🔢 Columns detected: {len(header_cols)}")
                    print(f"📋 Columns: {header_cols}")

                try:
                    read_kwargs = dict(
                        delimiter=delimiter, dtype=str, encoding='utf-8',
                        on_bad_lines='warn', quotechar='"',
                        skipinitialspace=True, keep_default_na=False
                    )
                    if has_headers:
                        df = pd.read_csv(file_path, **read_kwargs)
                    else:
                        df = pd.read_csv(file_path, header=None, **read_kwargs)
                except Exception:
                    read_kwargs['engine'] = 'python'
                    if has_headers:
                        df = pd.read_csv(file_path, **read_kwargs)
                    else:
                        df = pd.read_csv(file_path, header=None, **read_kwargs)
            else:
                df = pd.read_excel(file_path, dtype=str, keep_default_na=False)

            df.columns = [str(col).strip() for col in df.columns]
            df = df.fillna('')

            # Normalise last_outcome column
            last_outcome_col = None
            for col in df.columns:
                if 'last_outcome' in col.lower():
                    last_outcome_col = col
                    break
            if not last_outcome_col:
                for col in df.columns:
                    if any(t in col.lower() for t in ['outcome', 'result', 'status', 'disposition']):
                        last_outcome_col = col
                        break

            if last_outcome_col and last_outcome_col != 'last_outcome':
                df = df.rename(columns={last_outcome_col: 'last_outcome'})
            elif not last_outcome_col:
                df['last_outcome'] = ''

            # Normalise contact_id column
            contact_id_col = None
            for col in df.columns:
                if 'contact_id' in col.lower() or col.lower() == 'contact':
                    contact_id_col = col
                    break
            if contact_id_col and contact_id_col != 'contact_id':
                df = df.rename(columns={contact_id_col: 'contact_id'})
            elif not contact_id_col:
                df['contact_id'] = [f"ID_{i + 1}" for i in range(len(df))]

            # Build outcome map
            outcomes = OutcomeDescription.objects.all()
            outcome_map = {o.last_outcome: o.description for o in outcomes}
            print(f"📚 Loaded {len(outcome_map)} outcome descriptions")

            # Insert Description column right after last_outcome
            if 'last_outcome' in df.columns:
                col_idx = list(df.columns).index('last_outcome') + 1

                def get_description(outcome_code):
                    try:
                        if outcome_code is None or str(outcome_code).strip() == '':
                            return ''
                        value_str = str(outcome_code)
                        if value_str in outcome_map:
                            return outcome_map[value_str]
                        outcome = OutcomeDescription.objects.filter(
                            last_outcome=value_str
                        ).first()
                        if outcome:
                            outcome_map[value_str] = outcome.description
                            return outcome.description
                        outcome = OutcomeDescription.objects.filter(
                            last_outcome__iexact=value_str
                        ).first()
                        if outcome:
                            outcome_map[value_str] = outcome.description
                            return outcome.description
                        return value_str
                    except Exception as e:
                        print(f"❌ get_description error for '{outcome_code}': {e}")
                        return str(outcome_code) if outcome_code else ''

                df.insert(col_idx, 'Description', df['last_outcome'].apply(get_description))

            print(f"✅ Processed {len(df)} records")
            return df

        except Exception as e:
            print(f"❌ Error processing file {file_path}: {e}")
            traceback.print_exc()
            raise Exception(f"Error processing file {file_path}: {e}")


DataProcessor = SimpleDataProcessor


# ===========================================================
# CALL DATA FILE VIEWSET
# ===========================================================

class CallDataFileViewSet(viewsets.ModelViewSet):
    """Manage call data file uploads"""
    serializer_class = CallDataFileSerializer
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        queryset = CallDataFile.objects.all().order_by('-uploaded_at')
        campaign_id = self.request.query_params.get('campaign_id')
        if campaign_id:
            queryset = queryset.filter(campaign_id=campaign_id)
        return queryset

    def perform_create(self, serializer):
        serializer.save()

    @action(detail=True, methods=['get'])
    def preview(self, request, pk=None):
        """Preview processed data"""
        try:
            file_obj = self.get_object()
            processed_data = ProcessedData.objects.filter(call_data_file=file_obj)

            if not processed_data.exists():
                return Response(
                    {'success': False, 'error': 'No processed data found for this file'},
                    status=status.HTTP_404_NOT_FOUND
                )

            data = []
            for record in processed_data[:100]:
                row = {
                    'contact_id': record.contact_id,
                    'last_outcome': record.last_outcome,
                    'Description': record.outcome_description
                }
                data.append(row)

            return Response({
                'success': True,
                'data': {
                    'columns': list(data[0].keys()) if data else [],
                    'data': data,
                    'total_records': processed_data.count()
                }
            })
        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='download_processed')
    def download_processed(self, request, pk=None):
        """Download the processed version of a file as Excel"""
        try:
            file_obj = self.get_object()
            processed_data = ProcessedData.objects.filter(
                call_data_file=file_obj
            ).order_by('id')

            if not processed_data.exists():
                return Response(
                    {'success': False, 'error': 'No processed data found for this file.'},
                    status=status.HTTP_404_NOT_FOUND
                )

            field_names = [
                f.name for f in ProcessedData._meta.fields
                if f.name not in ['id', 'call_data_file', 'processed_at', 'outcome_description']
            ]

            data = []
            for record in processed_data:
                row = {}
                for field in field_names:
                    value = getattr(record, field, None)
                    if value is not None and value != '':
                        row[field] = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'strftime') else value
                    else:
                        row[field] = ''
                data.append(row)

            df = pd.DataFrame(data)
            descriptions = [record.outcome_description or '' for record in processed_data]

            columns = list(df.columns)
            last_outcome_pos = next(
                (i for i, col in enumerate(columns) if col == 'last_outcome'), -1
            )
            if last_outcome_pos != -1:
                df.insert(last_outcome_pos + 1, 'Description', descriptions)
            else:
                df['Description'] = descriptions

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Processed Data', index=False)
            output.seek(0)

            base_name = os.path.splitext(file_obj.original_name or 'data')[0]
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{base_name}_processed.xlsx"'
            return response

        except CallDataFile.DoesNotExist:
            return Response({'success': False, 'error': 'File not found'},
                            status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': f'Error generating file: {e}'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ===========================================================
# OUTCOME DESCRIPTION VIEWSET
# ===========================================================

class OutcomeDescriptionViewSet(viewsets.ModelViewSet):
    """Manage outcome descriptions"""
    queryset = OutcomeDescription.objects.all()
    serializer_class = OutcomeDescriptionSerializer
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        queryset = OutcomeDescription.objects.all()
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(last_outcome__icontains=search) | Q(description__icontains=search)
            )
        return queryset.order_by('last_outcome')

    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        serializer.save(created_by=user)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        """Bulk upload outcomes from Excel/CSV"""
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        try:
            if file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file, dtype=str, engine='openpyxl', keep_default_na=False)
            elif file.name.endswith('.csv'):
                df = pd.read_csv(file, dtype=str, encoding='utf-8',
                                 on_bad_lines='skip', keep_default_na=False)
            else:
                return Response({'error': 'Unsupported format. Use CSV or Excel.'},
                                status=status.HTTP_400_BAD_REQUEST)

            required_columns = ['last_outcome', 'Description']
            missing = [c for c in required_columns if c not in df.columns]
            if missing:
                return Response(
                    {'error': f'Missing columns: {", ".join(missing)}. Found: {list(df.columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            df = df.fillna('')
            df['last_outcome'] = df['last_outcome'].astype(str).str.strip()
            df['Description'] = df['Description'].astype(str).str.strip()

            initial_count = OutcomeDescription.objects.count()
            batch_size = 500
            created_count = 0
            errors = []
            outcomes_to_create = []

            for i in range(0, len(df), batch_size):
                batch = df.iloc[i:i + batch_size]
                for index, row in batch.iterrows():
                    try:
                        last_outcome = row['last_outcome']
                        description = row['Description']
                        if last_outcome == '' and description == '':
                            continue
                        outcomes_to_create.append(OutcomeDescription(
                            last_outcome=last_outcome,
                            description=description,
                            created_by=request.user if request.user.is_authenticated else None,
                            is_active=True
                        ))
                        created_count += 1
                    except Exception as e:
                        errors.append(f"Row {index + 2}: {e}")

                try:
                    OutcomeDescription.objects.bulk_create(
                        outcomes_to_create, batch_size=100, ignore_conflicts=True
                    )
                    outcomes_to_create = []
                except Exception as e:
                    for outcome in outcomes_to_create:
                        try:
                            outcome.save()
                        except IntegrityError:
                            pass
                        except Exception as e2:
                            errors.append(f"Save error: {e2}")
                    outcomes_to_create = []

            if outcomes_to_create:
                try:
                    OutcomeDescription.objects.bulk_create(
                        outcomes_to_create, batch_size=100, ignore_conflicts=True
                    )
                except Exception as e:
                    for outcome in outcomes_to_create:
                        try:
                            outcome.save()
                        except Exception:
                            pass

            final_count = OutcomeDescription.objects.count()
            response_data = {
                'message': f'Processed {len(df)} rows from file.',
                'details': {
                    'file_rows': len(df),
                    'processed_rows': created_count,
                    'initial_db_count': initial_count,
                    'final_db_count': final_count,
                    'total_added_to_db': final_count - initial_count,
                }
            }
            if errors:
                response_data['errors'] = errors[:20]
                response_data['error_count'] = len(errors)

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            traceback.print_exc()
            return Response({'error': f'Error processing file: {e}'},
                            status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export outcomes to Excel"""
        try:
            outcomes = OutcomeDescription.objects.all().order_by('last_outcome', 'id')
            data = [{
                'last_outcome': o.last_outcome,
                'Description': o.description,
                'Created By': o.created_by.username if o.created_by else '',
                'Created At': o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else '',
                'ID': o.id
            } for o in outcomes]

            df = pd.DataFrame(data)
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Outcome Descriptions', index=False)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="outcome_descriptions.xlsx"'
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_upload_outcomes(request):
    """Standalone bulk upload outcomes endpoint"""
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

    file = request.FILES['file']
    try:
        if file.name.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file, dtype=str, engine='openpyxl', keep_default_na=False)
        elif file.name.endswith('.csv'):
            df = pd.read_csv(file, dtype=str, encoding='utf-8',
                             on_bad_lines='skip', keep_default_na=False)
        else:
            return Response({'error': 'Unsupported format. Use CSV or Excel.'},
                            status=status.HTTP_400_BAD_REQUEST)

        required_columns = ['last_outcome', 'Description']
        missing = [c for c in required_columns if c not in df.columns]
        if missing:
            return Response(
                {'error': f'Missing columns: {", ".join(missing)}. Found: {list(df.columns)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        df = df.fillna('')
        df['last_outcome'] = df['last_outcome'].astype(str).str.strip()
        df['Description'] = df['Description'].astype(str).str.strip()

        initial_count = OutcomeDescription.objects.count()
        outcomes_to_create = []
        created_count = 0
        skipped_count = 0
        errors = []

        for index, row in df.iterrows():
            last_outcome = str(row['last_outcome']).strip()
            description = str(row['Description']).strip()
            if last_outcome == '' and description == '':
                skipped_count += 1
                continue
            outcomes_to_create.append(OutcomeDescription(
                last_outcome=last_outcome,
                description=description,
                created_by=request.user if request.user.is_authenticated else None,
                is_active=True
            ))
            created_count += 1

            if len(outcomes_to_create) >= 500:
                try:
                    OutcomeDescription.objects.bulk_create(
                        outcomes_to_create, batch_size=100, ignore_conflicts=True
                    )
                    outcomes_to_create = []
                except Exception as e:
                    for outcome in outcomes_to_create:
                        try:
                            outcome.save()
                        except Exception:
                            pass
                    outcomes_to_create = []

        if outcomes_to_create:
            try:
                OutcomeDescription.objects.bulk_create(
                    outcomes_to_create, batch_size=100, ignore_conflicts=True
                )
            except Exception:
                for outcome in outcomes_to_create:
                    try:
                        outcome.save()
                    except Exception:
                        pass

        final_count = OutcomeDescription.objects.count()
        return Response({
            'message': f'Processed {len(df)} rows.',
            'details': {
                'file_rows': len(df),
                'attempted_to_create': created_count,
                'skipped_empty_rows': skipped_count,
                'initial_db_count': initial_count,
                'final_db_count': final_count,
                'total_added_to_db': final_count - initial_count,
            }
        }, status=status.HTTP_200_OK)

    except Exception as e:
        traceback.print_exc()
        return Response({'error': f'Error processing file: {e}'},
                        status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([AllowAny])
def export_outcomes(request):
    """Export outcomes to Excel"""
    try:
        outcomes = OutcomeDescription.objects.all().order_by('last_outcome', 'id')
        data = [{
            'last_outcome': o.last_outcome,
            'Description': o.description,
            'Created By': o.created_by.username if o.created_by else '',
            'Created At': o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else '',
            'ID': o.id
        } for o in outcomes]

        df = pd.DataFrame(data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Outcome Descriptions', index=False)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="outcome_descriptions.xlsx"'
        return response
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ===========================================================
# REPORT VIEWSET
# ===========================================================

class ReportViewSet(
    viewsets.GenericViewSet,
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin
):
    """Generate and manage reports — campaign-scoped."""
    queryset = GeneratedReport.objects.all().order_by('-generated_at')
    serializer_class = GeneratedReportSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        queryset = GeneratedReport.objects.all().order_by('-generated_at')
        # FIX: filter by campaign when requested
        campaign_id = self.request.query_params.get('campaign_id')
        if campaign_id:
            queryset = queryset.filter(campaign_id=campaign_id)
        elif self.request.user.is_authenticated:
            queryset = queryset.filter(user=self.request.user)
        return queryset

    # ----------------------------------------------------------
    # HELPER: EXCEL FORMAT DEFINITIONS
    # ----------------------------------------------------------

    def _get_workbook_formats(self, workbook):
        return {
            'title': workbook.add_format({
                'bold': True, 'font_size': 18, 'align': 'center',
                'valign': 'vcenter', 'font_color': '#1F4E78'
            }),
            'header': workbook.add_format({
                'bold': True, 'bg_color': '#366092', 'font_color': 'white',
                'border': 1, 'align': 'center', 'valign': 'vcenter'
            }),
            'subheader': workbook.add_format({
                'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white',
                'border': 1, 'align': 'center', 'valign': 'vcenter'
            }),
            'cell': workbook.add_format({
                'border': 1, 'align': 'left', 'valign': 'vcenter'
            }),
            'number': workbook.add_format({
                'border': 1, 'align': 'center', 'valign': 'vcenter',
                'num_format': '#,##0'
            }),
            'percent': workbook.add_format({
                'border': 1, 'align': 'center', 'valign': 'vcenter',
                'num_format': '0.00%'
            }),
            'formula_cell': workbook.add_format({
                'border': 1, 'align': 'center', 'valign': 'vcenter',
                'num_format': '#,##0', 'bg_color': '#F2F2F2', 'bold': True
            }),
            'pivot_label': workbook.add_format({
                'bold': True, 'bg_color': '#D9E1F2', 'border': 1
            }),
            'grand_total': workbook.add_format({
                'bold': True, 'bg_color': '#8EA9DB', 'border': 1,
                'num_format': '#,##0'
            }),
            'summary_label': workbook.add_format({
                'bold': True, 'border': 1, 'align': 'left', 'valign': 'vcenter',
                'bg_color': '#E2EFDA'
            }),
            'summary_value': workbook.add_format({
                'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter',
                'num_format': '#,##0', 'bg_color': '#E2EFDA'
            }),
            'summary_percent': workbook.add_format({
                'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter',
                'num_format': '0.00%', 'bg_color': '#E2EFDA'
            }),
        }

    # ----------------------------------------------------------
    # DOWNLOAD
    # ----------------------------------------------------------

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """Download a generated report file."""
        try:
            report = self.get_object()
            if not report.file or not os.path.exists(report.file.path):
                return Response(
                    {'success': False, 'error': 'Report file not found on server.'},
                    status=status.HTTP_404_NOT_FOUND
                )
            with open(report.file.path, 'rb') as f:
                file_data = f.read()
            response = HttpResponse(
                file_data,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = os.path.basename(report.file.path)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except GeneratedReport.DoesNotExist:
            return Response({'success': False, 'error': 'Report not found'},
                            status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ----------------------------------------------------------
    # GENERATE CAMPAIGN REPORT  ← FIX: fully campaign-scoped
    # ----------------------------------------------------------

    @staticmethod
    def _auto_generate_full_report(file_instance):
        """
        Automatically called after a data file is processed.
        Generates ONE workbook with 4 sheets and saves it as a GeneratedReport:

            Sheet 1: Processed Data   — every record from this upload
            Sheet 2: Pivot            — count per outcome description
            Sheet 3: Campaign Analysis — summary metrics + category tables
            Sheet 4: Sheet1           — the campaign's template, populated

        This replaces the old two-step flow (Generate Report → Run Analysis).
        Everything is ready to download as soon as the upload completes.
        """
        from django.conf import settings
        import xlsxwriter, openpyxl, os, traceback
        from io import BytesIO
        from datetime import datetime
        from django.db.models import Count
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        campaign = file_instance.campaign
        if not campaign:
            print("⚠️  File has no campaign — skipping auto-report.")
            return

        print(f"\n{'='*60}")
        print(f"🚀 AUTO-REPORT: campaign='{campaign.display_name}' "
              f"file='{file_instance.original_name}'")

        # ── 1. Build outcome map ───────────────────────────────────────
        outcome_descriptions = OutcomeDescription.objects.all()
        outcome_map = {o.last_outcome: o.description for o in outcome_descriptions}
        print(f"📚 Outcome map: {len(outcome_map)} entries")

        # ── 2. Load processed data for this file ───────────────────────
        processed_data_query = ProcessedData.objects.filter(
            call_data_file=file_instance
        )
        total_count = processed_data_query.count()
        print(f"📊 Processed records: {total_count}")

        if total_count == 0:
            print("⚠️  No processed data — skipping auto-report.")
            return

        # ── 3. Build Pivot counts ──────────────────────────────────────
        raw_counts = processed_data_query.values('last_outcome').annotate(
            count=Count('id')
        )
        description_counts = {}
        grand_total = 0
        for item in raw_counts:
            key  = item['last_outcome'] or 'Unknown'
            desc = outcome_map.get(key, key)
            description_counts[desc] = description_counts.get(desc, 0) + item['count']
            grand_total += item['count']

        sorted_desc_counts = sorted(
            description_counts.items(), key=lambda x: x[1], reverse=True
        )

        # ── 4. Compute summary metrics ─────────────────────────────────
        total_leads = grand_total
        unworked_leads = processed_data_query.filter(
            last_outcome__iexact='New'
        ).count()

        SALE_TERMS = [
            'sale made', 'upsell', 'tyme bank account sale',
            'sale made - completed mandate', 'sale made - pending mandate',
        ]
        TRUE_CONTACT_TERMS = [
            'not interested', 'callback', 'call back', 'client hung up',
            'affordability', 'cannot afford', 'declined sale', 'qa rework',
            'qa fail', 'quoted client not interested', 'not interested upfront',
            'call back hold', 'not interested transaction fees',
            'cannot afford premium', 'not interested household contents',
            'not interested sms', 'call back via ms teams',
        ]

        true_sales = 0
        true_contacts = 0
        unworked_leads = 0

        for desc, count in description_counts.items():
            d = desc.lower().strip()
            if any(t in d for t in SALE_TERMS):
                true_sales += count
            elif any(t in d for t in TRUE_CONTACT_TERMS):
                true_contacts += count
            elif any(t == d or d.startswith(t) for t in ['new', 'not contacted']):
                unworked_leads += count

        successful_contacts = true_contacts + true_sales
        conversion_value    = (true_sales / total_leads * 100) if total_leads > 0 else 0
        conversion_decimal  = conversion_value / 100

        print(f"📊 Metrics: TL={total_leads} SC={successful_contacts} "
              f"TC={true_contacts} TS={true_sales} Conv={conversion_value:.2f}%")

        # ── 5. Build workbook ──────────────────────────────────────────
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'nan_inf_to_errors': True})

        # Formats
        fmts = {}
        fmts['title'] = workbook.add_format({
            'bold': True, 'font_size': 18, 'align': 'center',
            'valign': 'vcenter', 'font_color': '#1F4E78'
        })
        fmts['header'] = workbook.add_format({
            'bold': True, 'bg_color': '#366092', 'font_color': 'white',
            'border': 1, 'align': 'center', 'valign': 'vcenter'
        })
        fmts['subheader'] = workbook.add_format({
            'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white',
            'border': 1, 'align': 'center', 'valign': 'vcenter'
        })
        fmts['cell'] = workbook.add_format({
            'border': 1, 'align': 'left', 'valign': 'vcenter'
        })
        fmts['number'] = workbook.add_format({
            'border': 1, 'align': 'center', 'num_format': '#,##0'
        })
        fmts['percent'] = workbook.add_format({
            'border': 1, 'align': 'center', 'num_format': '0.00%'
        })
        fmts['formula_cell'] = workbook.add_format({
            'border': 1, 'align': 'center', 'num_format': '#,##0',
            'bg_color': '#F2F2F2', 'bold': True
        })
        fmts['grand_total'] = workbook.add_format({
            'bold': True, 'bg_color': '#8EA9DB', 'border': 1,
            'num_format': '#,##0'
        })
        fmts['summary_label'] = workbook.add_format({
            'bold': True, 'border': 1, 'align': 'left',
            'bg_color': '#E2EFDA'
        })
        fmts['summary_value'] = workbook.add_format({
            'bold': True, 'border': 1, 'align': 'center',
            'num_format': '#,##0', 'bg_color': '#E2EFDA'
        })
        fmts['summary_percent'] = workbook.add_format({
            'bold': True, 'border': 1, 'align': 'center',
            'num_format': '0.00%', 'bg_color': '#E2EFDA'
        })
        fmts['ca_title'] = workbook.add_format({
            'bold': True, 'font_size': 16, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#D9D9D9', 'font_color': '#000000', 'border': 1
        })
        fmts['ca_super_header'] = workbook.add_format({
            'bold': True, 'align': 'center', 'valign': 'vcenter',
            'bg_color': '#D9D9D9', 'font_color': '#000000', 'border': 1
        })
        fmts['ca_total_number'] = workbook.add_format({
            'bold': True, 'font_size': 32, 'align': 'center', 'valign': 'vcenter',
            'border': 1, 'num_format': '#,##0'
        })
        fmts['ca_italic_note'] = workbook.add_format({
            'italic': True, 'align': 'left', 'valign': 'vcenter', 'border': 1
        })
        fmts['ca_grand_total'] = workbook.add_format({
            'bold': True, 'font_size': 14, 'align': 'left', 'valign': 'vcenter'
        })

        # ── SHEET 1: PROCESSED DATA ────────────────────────────────────
        data_ws = workbook.add_worksheet('Processed Data')
        all_records = list(processed_data_query[:10000])
        field_names = [
            f.name for f in ProcessedData._meta.fields
            if f.name not in ['id', 'call_data_file', 'processed_at']
        ]
        for col, field in enumerate(field_names):
            data_ws.write(0, col, field.replace('_', ' ').title(), fmts['header'])
        for row_num, record in enumerate(all_records, start=1):
            for col, field in enumerate(field_names):
                val = getattr(record, field, '')
                if val is not None and val != '':
                    data_ws.write(
                        row_num, col,
                        val.strftime('%Y-%m-%d %H:%M:%S') if hasattr(val, 'strftime') else val
                    )
        print(f"✅ Processed Data: {len(all_records)} rows")

        # ── SHEET 2: PIVOT ─────────────────────────────────────────────
        pivot_ws = workbook.add_worksheet('Pivot')
        pivot_ws.set_column('A:A', 40)
        pivot_ws.set_column('B:B', 20)
        pivot_ws.write('A1', 'Outcome Description', fmts['header'])
        pivot_ws.write('B1', 'Count', fmts['header'])
        curr_row = 1
        for desc, count in sorted_desc_counts:
            pivot_ws.write(curr_row, 0, desc, fmts['cell'])
            pivot_ws.write(curr_row, 1, count, fmts['number'])
            curr_row += 1
        pivot_ws.write(curr_row, 0, 'Grand Total', fmts['grand_total'])
        pivot_ws.write(curr_row, 1, grand_total, fmts['grand_total'])
        pivot_range = f"Pivot!$A$2:$B${curr_row}"
        print(f"✅ Pivot: {curr_row - 1} unique outcomes")

        # ── SHEET 3: CAMPAIGN ANALYSIS ─────────────────────────────────
        ca_ws = workbook.add_worksheet('Campaign Analysis')
        for i, w in enumerate([18,32,12,32,12,32,12,32,12]):
            ca_ws.set_column(i, i, w)

        title_suffix = 'Campaign Analysis' if campaign.display_name.strip().lower().endswith('leads') \
            else 'Leads Campaign Analysis'
        ca_ws.merge_range(
            'A1:I2',
            f'{campaign.display_name} {title_suffix}',
            fmts['ca_title']
        )

        categories = {
            'unsuccessful': [
                'Answering Machine Autodial','No Answer Autodial','Auto Engaged',
                'Disconnected Number Auto','Answering Machine','Selected','No Answer',
                'Busy Tone','Dropped','New','Call Dropped','Inbound After Hours Drop',
                'TPS Registered Number','Temporary Disconnected Number','Bad Line Quality',
                'Inbound Abandon','Outbound Pre-Routing Drop','Engaged','Disconnected',
                'CONGESTED','Flow Inbound Abandon','Multiple Calls','Missed','Answered',
                'Customer drop','Auto Dial Disconnected','Disconnected Number Temporary'
            ],
            'successful': [
                'Sale Made - Completed Mandate','Sale Made - Pending Mandate',
                'Sale Made','Tyme Bank Account Sale','QA Rework'
            ],
            'unworkable': [
                'Already Contacted','Existing Client','Unemployed','Wrong Number',
                'Do Not Call','No Smartphones','Client Deceased',
                'Right Party Not Available','Client Overage Limit','Language Barrier',
                'Non SA Citizen','No Bank Account','Client Underage','Go To The Branch',
                'Completed','Cannot Afford Upfront Payment','Account Suspended',
                'Insured at another company','Scheduled Appointment','Refund Request',
                'Does Not Qualify','NTU Policy','Does Not Have a Business',
                "Refer To Store - Doesn't want to complete online",
                'Does Not Need It Now','Technical Issue','Unsuccessful Application'
            ],
            'true_contacts': [
                'Client Hung Up','CallBack','Not Interested','Not Interested Upfront',
                'Cannot Afford Premium','Not interested - Pitched','Affordability',
                'Quoted Client Not Interested',
                'Not interested business does not use speed point mac',
                'Call Back Technical Error','Not Interested Transaction Fees',
                'Cannot Afford','Not Interested Household Contents',
                'Call Back Hold','Declined Sale'
            ]
        }

        # Super-header row: "Customers Reached" spans the True Contacts pair only
        SUPER_HDR = 2
        ca_ws.merge_range(SUPER_HDR, 7, SUPER_HDR, 8, 'Customers Reached', fmts['ca_super_header'])

        HEADER_ROW = 3
        ca_ws.write(HEADER_ROW, 0, 'Total Leads Dialled',              fmts['header'])
        ca_ws.write(HEADER_ROW, 1, 'Unsuccessful Contacts',            fmts['header'])
        ca_ws.write(HEADER_ROW, 2, 'Lead Count',                       fmts['header'])
        ca_ws.write(HEADER_ROW, 3, 'Was customer interested in deal?', fmts['header'])
        ca_ws.write(HEADER_ROW, 4, 'Lead Count',                       fmts['header'])
        ca_ws.write(HEADER_ROW, 5, 'Succesful Contacts',                fmts['header'])
        ca_ws.write(HEADER_ROW, 6, 'Lead Count',                       fmts['header'])
        ca_ws.write(HEADER_ROW, 7, 'True Contacts',                    fmts['header'])
        ca_ws.write(HEADER_ROW, 8, 'Lead Count',                       fmts['header'])

        DATA_START = HEADER_ROW + 1

        def fill_section(desc_list, col_label, col_val, start_row):
            for i, text in enumerate(desc_list):
                r = start_row + i
                ca_ws.write(r, col_label, text, fmts['cell'])
                ca_ws.write_formula(
                    r, col_val,
                    f'=IFERROR(VLOOKUP("{text}",{pivot_range},2,FALSE),0)',
                    fmts['formula_cell']
                )
            return start_row + len(desc_list)

        u_end = fill_section(categories['unsuccessful'], 1, 2, DATA_START)
        s_end = fill_section(categories['successful'],   3, 4, DATA_START)
        w_end = fill_section(categories['unworkable'],   5, 6, DATA_START)
        t_end = fill_section(categories['true_contacts'],7, 8, DATA_START)

        # The two longest columns (Unsuccessful / Succesful Contacts) set where every
        # column's subtotal row sits.
        SUBTOTAL_ROW = DATA_START + max(
            len(categories['unsuccessful']), len(categories['unworkable'])
        )

        # "If not, why not?" note fills the gap between the 5 sale rows and the subtotal
        # row, holding the count of everyone who was reached but didn't buy.
        gap_start, gap_end = s_end, SUBTOTAL_ROW - 1
        why_not_formula = f'=G{SUBTOTAL_ROW+1}+I{SUBTOTAL_ROW+1}'
        if gap_end > gap_start:
            ca_ws.merge_range(gap_start, 3, gap_end, 3, 'If not, why not?', fmts['ca_italic_note'])
            ca_ws.merge_range(gap_start, 4, gap_end, 4, why_not_formula, fmts['formula_cell'])
        elif gap_end == gap_start:
            ca_ws.write(gap_start, 3, 'If not, why not?', fmts['ca_italic_note'])
            ca_ws.write_formula(gap_start, 4, why_not_formula, fmts['formula_cell'])

        # Subtotal row — one number per category, aligned under the longest columns
        ca_ws.write_formula(SUBTOTAL_ROW, 2, f'=SUM(C{DATA_START+1}:C{u_end})', fmts['formula_cell'])
        ca_ws.write(SUBTOTAL_ROW, 3, "Successful Take Up's", fmts['subheader'])
        ca_ws.write_formula(SUBTOTAL_ROW, 4, f'=SUM(E{DATA_START+1}:E{s_end})', fmts['formula_cell'])
        ca_ws.write_formula(SUBTOTAL_ROW, 6, f'=SUM(G{DATA_START+1}:G{w_end})', fmts['formula_cell'])
        ca_ws.write_formula(SUBTOTAL_ROW, 8, f'=SUM(I{DATA_START+1}:I{t_end})', fmts['formula_cell'])

        # Big "Total Leads Dialled" number spans the full height of the data + subtotal rows
        ca_ws.merge_range(DATA_START, 0, SUBTOTAL_ROW, 0, grand_total, fmts['ca_total_number'])

        # Grand total of the "successful pipeline": Take Ups + Succesful Contacts + True Contacts
        GRAND_ROW = SUBTOTAL_ROW + 2
        ca_ws.write_formula(
            GRAND_ROW, 0,
            f'=E{SUBTOTAL_ROW+1}+G{SUBTOTAL_ROW+1}+I{SUBTOTAL_ROW+1}',
            fmts['ca_grand_total']
        )

        # ── SUMMARY + ranked status tables ─────────────────────────────
        def top_n_from_category(cat_list, n):
            """Top n (description, count) pairs from sorted_desc_counts that
            belong to the given category list, case-insensitive."""
            cat_lower = {c.lower() for c in cat_list}
            matched = [(d, c) for d, c in sorted_desc_counts if d.lower() in cat_lower]
            return matched[:n]

        SUMMARY_ROW = GRAND_ROW + 3
        ca_ws.merge_range(SUMMARY_ROW, 0, SUMMARY_ROW, 1, 'SUMMARY', fmts['subheader'])
        conversion_ratio = (true_sales / true_contacts) if true_contacts else 0
        contact_ratio = (successful_contacts / grand_total) if grand_total else 0
        summary_items = [
            ('Successful Contacts',      successful_contacts, fmts['formula_cell']),
            ('True Contacts (TC)',       true_contacts,        fmts['formula_cell']),
            ('Conversions',              true_sales,           fmts['formula_cell']),
            ('Conversion Ratio (vs TC)', conversion_ratio,     fmts['percent']),
            ('Contact Ratio',            contact_ratio,        fmts['percent']),
        ]
        for i, (label, value, fmt) in enumerate(summary_items):
            r = SUMMARY_ROW + 1 + i
            ca_ws.write(r, 0, label, fmts['cell'])
            ca_ws.write(r, 1, value, fmt)

        TOTAL_LEADS_ROW = SUMMARY_ROW + len(summary_items) + 2
        ca_ws.write(TOTAL_LEADS_ROW, 0, 'TOTAL LEADS', fmts['subheader'])
        ca_ws.write(TOTAL_LEADS_ROW, 1, grand_total, fmts['formula_cell'])

        def write_ranked_table(title, rows_data, start_row):
            ca_ws.merge_range(start_row, 0, start_row, 2, title, fmts['subheader'])
            ca_ws.write(start_row + 1, 0, 'Status',        fmts['header'])
            ca_ws.write(start_row + 1, 1, 'Lead Count',    fmts['header'])
            ca_ws.write(start_row + 1, 2, 'Lead % Result', fmts['header'])
            r = start_row + 2
            total_count = 0
            for desc, count in rows_data:
                ca_ws.write(r, 0, desc, fmts['cell'])
                ca_ws.write(r, 1, count, fmts['number'])
                ca_ws.write(r, 2, (count / grand_total) if grand_total else 0, fmts['percent'])
                total_count += count
                r += 1
            ca_ws.write(r, 0, 'TOTAL', fmts['subheader'])
            ca_ws.write(r, 1, total_count, fmts['formula_cell'])
            ca_ws.write(r, 2, (total_count / grand_total) if grand_total else 0, fmts['percent'])
            return r + 2  # next block starts 1 blank row below

        next_row = write_ranked_table(
            'Top 5 Failed Status Codes',
            top_n_from_category(categories['unsuccessful'], 5),
            TOTAL_LEADS_ROW + 2
        )
        write_ranked_table(
            'Successful Leads',
            top_n_from_category(categories['successful'], 3),
            next_row
        )

        print(f"✅ Campaign Analysis sheet built")

        # ── SHEET 4: TEMPLATE (Sheet1) populated from Pivot ────────────
        # Find the newest template for this campaign
        template_obj = ReportTemplate.objects.filter(
            campaign=campaign, is_active=True
        ).order_by('-uploaded_at').first()

        if template_obj and os.path.exists(template_obj.template_file.path):
            print(f"📋 Populating template: {template_obj.name}")

            # Save and re-open the workbook so Pivot data is readable
            # by openpyxl (xlsxwriter can't be read while open)
            workbook.close()
            output.seek(0)
            xl_wb = load_workbook(output)
            xl_pivot = xl_wb['Pivot']

            # ══════════════════════════════════════════════════════════
            # Build Sheet1's lookup FROM THE CAMPAIGN ANALYSIS SHEET DATA
            # ══════════════════════════════════════════════════════════
            # Sheet1 mirrors the Campaign Analysis sheet exactly:
            #
            #   1. The SUMMARY BLOCK values (Total Leads, Successful
            #      Contacts, True Contacts, Conversion, True Sales,
            #      Unworked Leads) — the same numbers written to the
            #      Campaign Analysis summary rows.
            #
            #   2. EVERY disposition listed in the Campaign Analysis
            #      category tables (Unsuccessful / Interested in Deal /
            #      Successful / True Contacts columns) — with the same
            #      count its VLOOKUP resolves to. Dispositions with no
            #      Pivot entry get 0, exactly like IFERROR(VLOOKUP,0)
            #      shows 0 in the Campaign Analysis sheet.
            #
            # (The Campaign Analysis sheet's Lead Count cells are live
            #  VLOOKUP formulas that Excel hasn't calculated yet, so we
            #  can't literally read them with openpyxl — instead we use
            #  the SAME source data that produces them, which guarantees
            #  identical values.)

            # Case-insensitive view of the Pivot counts
            desc_counts_lower = {
                k.strip().lower(): v for k, v in description_counts.items()
            }

            pivot_data_by_description = {}

            # 2a. Every disposition in the Campaign Analysis category
            #     tables — count mirrors the CA sheet's VLOOKUP result
            for cat_list in categories.values():
                for desc in cat_list:
                    key   = desc.strip().lower()
                    count = desc_counts_lower.get(key, 0)   # IFERROR(...,0)
                    pivot_data_by_description[key] = [desc, count]

            # 2b. Also include any Pivot description NOT in the category
            #     lists, so unusual dispositions still match in Sheet1
            for i, row in enumerate(xl_pivot.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if row and row[0]:
                    key = str(row[0]).strip().lower()
                    if key not in pivot_data_by_description and key != 'grand total':
                        pivot_data_by_description[key] = list(row)

            # 1. SUMMARY BLOCK values — identical to what the Campaign
            #    Analysis sheet's summary rows show
            summary_entries = {
                'total leads':                 ('total leads',                total_leads),
                'unworked leads (new leads)':  ('unworked leads (new leads)', unworked_leads),
                'succesful contacts':           ('succesful contacts',         successful_contacts),
                'successful contacts':          ('successful contacts',        successful_contacts),
                'true contacts':               ('true contacts',               true_contacts),
                'conversion':                   ('conversion',                 conversion_decimal),
                'true sales (post qa)':        ('true sales (post qa)',        true_sales),
            }
            # Calculated ratios
            if total_leads > 0:
                summary_entries['contactability']         = ('contactability',         successful_contacts / total_leads)
                summary_entries['lead to sale conversion'] = ('lead to sale conversion', true_sales / total_leads)
            if successful_contacts > 0:
                summary_entries['conversion to true sale'] = ('conversion to true sale', true_sales / successful_contacts)
            if true_contacts > 0:
                summary_entries['true contacts to sales']  = ('true contacts to sales',  true_sales / true_contacts)

            # Aliases for common template typos/name differences
            ALIASES = {
                'succesful contacts':        'successful contacts',
                'under age':                 'client underage',
                'go to branch':              'go to the branch',
                'qa fail':                   'qa rework',
                'upsell':                    'tyme bank account sale',
                'not working cannot afford': 'cannot afford',
            }
            for alias, real_key in ALIASES.items():
                if real_key in pivot_data_by_description and alias not in pivot_data_by_description:
                    pivot_data_by_description[alias] = pivot_data_by_description[real_key]

            for label, row_data in summary_entries.items():
                pivot_data_by_description[label] = list(row_data)

            # Load template and recreate with values
            src_wb    = load_workbook(template_obj.template_file.path)
            new_wb    = openpyxl.Workbook()
            new_wb.remove(new_wb.active)

            for sheet_name in src_wb.sheetnames:
                src_sheet = src_wb[sheet_name]
                new_sheet = new_wb.create_sheet(title=sheet_name)

                for row in src_sheet.iter_rows():
                    for cell in row:
                        nc = new_sheet.cell(row=cell.row, column=cell.column)
                        if cell.has_style:
                            try:
                                nc.font          = cell.font.copy()
                                nc.border        = cell.border.copy()
                                nc.fill          = cell.fill.copy()
                                nc.number_format = cell.number_format
                                nc.alignment     = cell.alignment.copy()
                            except Exception:
                                pass
                        # Preserve formulas; keep col-1 labels; clear everything else
                        if cell.data_type == 'f' and cell.value and str(cell.value).startswith('='):
                            nc.value = cell.value
                        elif cell.column == 1:
                            nc.value = cell.value
                        else:
                            nc.value = None

                for mr in src_sheet.merged_cells.ranges:
                    new_sheet.merge_cells(str(mr))
                for col_idx in range(1, src_sheet.max_column + 1):
                    cl = get_column_letter(col_idx)
                    if cl in src_sheet.column_dimensions:
                        new_sheet.column_dimensions[cl].width = src_sheet.column_dimensions[cl].width
                for ri in range(1, src_sheet.max_row + 1):
                    if ri in src_sheet.row_dimensions:
                        new_sheet.row_dimensions[ri].height = src_sheet.row_dimensions[ri].height

            # Populate matching rows in the first sheet of the template.
            #
            # RULES (per user requirements):
            #   - NEVER touch column A — the template's labels stay exactly
            #     as uploaded (no overwriting, no case changes)
            #   - Write ONLY the VALUE, into the template's real data column.
            #     Templates often merge column A across many columns (e.g.
            #     A:JK merged, data at JL) — so the data column is detected
            #     as (widest col-A merge end) + 1, falling back to col 2.
            #   - If the target cell is inside another merge, redirect the
            #     write to that merge's master (top-left) cell.
            #   - Formula cells are never overwritten.
            target_sheet_name = src_wb.sheetnames[0]
            tws = new_wb[target_sheet_name]
            rows_populated = 0

            # Detect the data column from column-A merges
            a_merge_end = 1
            for mr in tws.merged_cells.ranges:
                if mr.min_col == 1 and mr.max_col > a_merge_end:
                    a_merge_end = mr.max_col
            data_col = a_merge_end + 1 if a_merge_end > 1 else 2
            print(f"📌 Sheet1 data column: {data_col} "
                  f"({get_column_letter(data_col)}) — "
                  f"col A merges end at {a_merge_end}")

            # Merge-master lookup so writes inside merges go to the
            # top-left (writable) cell of that merge
            merge_master = {}
            for mr in tws.merged_cells.ranges:
                master = (mr.min_row, mr.min_col)
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        merge_master[(r, c)] = master

            for row_idx in range(1, tws.max_row + 1):
                desc_cell = tws.cell(row=row_idx, column=1)
                desc = str(desc_cell.value).strip().lower() if desc_cell.value else ""
                if not desc:
                    continue
                if desc not in pivot_data_by_description:
                    continue

                row_data = pivot_data_by_description[desc]
                # row_data shape: [description, value] — we only want the VALUE
                val = row_data[1] if len(row_data) > 1 else None
                if val is None:
                    continue

                # Resolve merge master for the target cell
                write_row, write_col = merge_master.get(
                    (row_idx, data_col), (row_idx, data_col)
                )
                cell = tws.cell(row=write_row, column=write_col)

                # Never overwrite formulas
                if cell.data_type == 'f' or (
                    cell.value is not None
                    and isinstance(cell.value, str)
                    and cell.value.strip().startswith('=')
                ):
                    continue

                try:
                    if isinstance(val, float) and val != int(val):
                        cell.value = float(val)     # ratios/percentages
                    elif isinstance(val, (int, float)):
                        cell.value = int(val)
                    else:
                        cell.value = val
                except Exception:
                    cell.value = val

                rows_populated += 1
                print(f"  ✅ row {row_idx:3d} → "
                      f"{get_column_letter(write_col)}{write_row}: "
                      f"'{desc_cell.value}' = {val}")

            print(f"✅ Sheet1 populated: {rows_populated} rows matched "
                  f"(labels untouched, values in col "
                  f"{get_column_letter(data_col)})")

            # Add the 3 existing sheets (Processed Data, Pivot, Campaign Analysis)
            # from the xlsxwriter output into the new_wb
            for extra_name in ['Processed Data', 'Pivot', 'Campaign Analysis']:
                if extra_name in xl_wb.sheetnames:
                    src_extra = xl_wb[extra_name]
                    dst_extra = new_wb.create_sheet(title=extra_name)
                    for row in src_extra.iter_rows():
                        for cell in row:
                            nc = dst_extra.cell(row=cell.row, column=cell.column)
                            nc.value = cell.value
                            if cell.has_style:
                                try:
                                    nc.font = cell.font.copy()
                                    nc.border = cell.border.copy()
                                    nc.fill = cell.fill.copy()
                                    nc.number_format = cell.number_format
                                    nc.alignment = cell.alignment.copy()
                                except Exception:
                                    pass
                    for mr in src_extra.merged_cells.ranges:
                        dst_extra.merge_cells(str(mr))
                    # Column widths / row heights aren't cell properties — copy them
                    # separately, or wide merged headers (e.g. Campaign Analysis) come
                    # out at Excel's default width and truncate.
                    for col_letter, dim in src_extra.column_dimensions.items():
                        if dim.width:
                            dst_extra.column_dimensions[col_letter].width = dim.width
                    for row_idx, dim in src_extra.row_dimensions.items():
                        if dim.height:
                            dst_extra.row_dimensions[row_idx].height = dim.height

            # Reorder sheets EXACTLY as requested:
            # 1. Processed Data  2. Sheet1  3. Campaign Analysis  4. Pivot
            desired_order = ['Processed Data', target_sheet_name, 'Campaign Analysis', 'Pivot']
            for i, name in enumerate(desired_order):
                if name in new_wb.sheetnames:
                    idx = new_wb.sheetnames.index(name)
                    new_wb.move_sheet(name, offset=i - idx)

            # Save final combined workbook
            final_output = BytesIO()
            new_wb.save(final_output)
            final_output.seek(0)
            file_bytes = final_output.getvalue()

        else:
            # No template — save the 3-sheet workbook (no Sheet1)
            print(f"⚠️  No template found for campaign '{campaign.display_name}' "
                  f"— saving 3-sheet report (no Sheet1)")
            workbook.close()
            output.seek(0)
            file_bytes = output.getvalue()

        # ── Save the report to disk ────────────────────────────────────
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename  = f"full_report_{campaign.name}_{timestamp}.xlsx"
        filepath  = os.path.join(reports_dir, filename)

        with open(filepath, 'wb') as f:
            f.write(file_bytes)

        # ── Save GeneratedReport record ────────────────────────────────
        report = GeneratedReport.objects.create(
            user=file_instance.user,
            campaign=campaign,
            report_type='campaign_analysis',
            file=f"reports/{filename}",
            parameters={
                'campaign_id':     campaign.id,
                'campaign_name':   campaign.display_name,
                'source_file':     file_instance.original_name,
                'record_count':    grand_total,
                'auto_generated':  True,
                'has_sheet1':      template_obj is not None,
                'template_name':   template_obj.name if template_obj else None,
                'rows_populated':  rows_populated if template_obj else 0,
                'metrics': {
                    'total_leads':         total_leads,
                    'unworked_leads':      unworked_leads,
                    'successful_contacts': successful_contacts,
                    'true_contacts':       true_contacts,
                    'conversion_pct':      round(conversion_value, 2),
                    'true_sales':          true_sales,
                }
            }
        )

        print(f"✅ Report saved: {filename} (ID: {report.id})")
        print(f"📑 Sheets: {new_wb.sheetnames if template_obj else '3-sheet (no template)'}")
        print(f"{'='*60}\n")
        return report

    @action(detail=False, methods=['post'])
    def generate_campaign(self, request):
        """
        Generate the FULL 4-sheet campaign report on demand.

        Produces ONE workbook with sheets in this exact order:
          1. Processed Data     – raw records from the latest upload
          2. Sheet1 (template)  – the campaign's template, populated with values
          3. Campaign Analysis  – summary metrics + category breakdowns
          4. Pivot              – counts per outcome description

        This is the SAME file that gets auto-generated when a data file is
        uploaded — this endpoint just regenerates it on demand (e.g. after
        uploading a new template or re-uploading data).

        Required POST body field: campaign_id
        """
        try:
            print("=" * 50)
            print("📊 GENERATE REPORT (manual trigger)")

            campaign_id = request.data.get('campaign_id')
            if not campaign_id:
                return Response(
                    {'success': False,
                     'error': 'campaign_id is required in the request body.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            try:
                campaign_obj = Campaign.objects.get(id=campaign_id, is_active=True)
            except Campaign.DoesNotExist:
                return Response(
                    {'success': False,
                     'error': f'Campaign with ID {campaign_id} not found.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            user = request.user if request.user.is_authenticated else None

            # Find the latest processed file for THIS campaign
            qs = CallDataFile.objects.filter(
                campaign=campaign_obj,
                status='processed'
            )
            if user:
                qs = qs.filter(user=user)
            latest_file = qs.order_by('-uploaded_at').first()

            if not latest_file:
                return Response(
                    {'success': False,
                     'error': f'No processed files found for campaign '
                              f'"{campaign_obj.display_name}". '
                              f'Upload a data file first.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            print(f"📁 Campaign : {campaign_obj.display_name}")
            print(f"📁 Source   : {latest_file.original_name}")

            # Delegate to the SAME generator used by auto-generation,
            # so manual and automatic reports are always identical.
            report = ReportViewSet._auto_generate_full_report(latest_file)

            if report is None:
                return Response(
                    {'success': False,
                     'error': 'Report generation produced no output. '
                              'Check that the file has processed records.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            metrics = report.parameters.get('metrics', {})

            return Response({
                'success': True,
                'data': {
                    'report_id':    report.id,
                    'download_url': f'/api/reports/{report.id}/download/',
                    'campaign':     campaign_obj.display_name,
                    'record_count': report.parameters.get('record_count', 0),
                    'has_sheet1':   report.parameters.get('has_sheet1', False),
                    'metrics':      metrics,
                    'message':      'Full 4-sheet report generated successfully.',
                }
            })

        except Exception as e:
            print(f"❌ REPORT ERROR: {e}")
            traceback.print_exc()
            return Response(
                {'success': False, 'error': f'Generation failed: {e}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='debug_analysis')
    def debug_analysis(self, request):
        """
        Debug endpoint: shows what's in the Pivot vs what's in the template.
        Call: GET /api/reports/debug_analysis/?campaign_id=X&template_id=Y
        """
        campaign_id = request.query_params.get('campaign_id')
        template_id = request.query_params.get('template_id')

        result = {}

        # Show Pivot content
        if campaign_id:
            report = GeneratedReport.objects.filter(
                report_type='campaign_analysis',
                campaign_id=campaign_id
            ).order_by('-generated_at').first()

            if report:
                try:
                    wb = load_workbook(report.file.path)
                    result['campaign_report_id'] = report.id
                    result['campaign_report_sheets'] = wb.sheetnames

                    if 'Pivot' in wb.sheetnames:
                        pivot_ws = wb['Pivot']
                        pivot_rows = []
                        for i, row in enumerate(pivot_ws.iter_rows(values_only=True)):
                            if i > 50: break  # limit output
                            pivot_rows.append({'col_a': row[0], 'col_b': row[1]})
                        result['pivot_rows'] = pivot_rows
                        result['pivot_row_count'] = pivot_ws.max_row
                    else:
                        result['pivot_error'] = 'No Pivot sheet found'
                except Exception as e:
                    result['campaign_report_error'] = str(e)
            else:
                result['campaign_report_error'] = f'No campaign_analysis report found for campaign {campaign_id}'

        # Show template content
        if template_id:
            try:
                template = ReportTemplate.objects.get(id=template_id)
                result['template_name'] = template.name
                result['template_sheets'] = template.sheet_names

                wb2 = load_workbook(template.template_file.path)
                result['template_actual_sheets'] = wb2.sheetnames

                # Show column A of first sheet
                first_sheet = wb2.active
                sheet_labels = []
                for row_idx in range(1, min(first_sheet.max_row + 1, 100)):
                    cell_a = first_sheet.cell(row=row_idx, column=1).value
                    cell_b = first_sheet.cell(row=row_idx, column=2).value
                    sheet_labels.append({
                        'row': row_idx,
                        'col_a': str(cell_a) if cell_a is not None else None,
                        'col_b_type': 'formula' if (cell_b and str(cell_b).startswith('=')) else 'value',
                        'col_b_val': str(cell_b)[:50] if cell_b is not None else None,
                    })
                result['template_first_sheet_labels'] = sheet_labels
            except Exception as e:
                result['template_error'] = str(e)

        return Response(result)

    @action(detail=False, methods=['post'], url_path='generate_campaign_analysis')
    def generate_campaign_analysis(self, request):
        """
        Generate a template-based campaign analysis report.

        Required POST body fields:
          - template_id   : ID of the uploaded ReportTemplate
          - campaign_name : Name of the sheet to populate in the template
          - campaign_id   : ID of the Campaign (for scoping)
        """
        try:
            template_id   = request.data.get('template_id')
            campaign_name = request.data.get('campaign_name')
            campaign_id   = request.data.get('campaign_id')

            print(f"📊 generate_campaign_analysis: template={template_id}, "
                  f"sheet={campaign_name}, campaign_id={campaign_id}")

            if not template_id:
                return Response({'success': False, 'error': 'template_id is required'},
                                status=status.HTTP_400_BAD_REQUEST)
            if not campaign_name:
                return Response({'success': False, 'error': 'campaign_name is required'},
                                status=status.HTTP_400_BAD_REQUEST)
            if not campaign_id:
                return Response({'success': False, 'error': 'campaign_id is required'},
                                status=status.HTTP_400_BAD_REQUEST)

            result = TemplateBasedReportGenerator.generate_analysis_report(
                template_id=template_id,
                data_mappings={'campaign': campaign_name},
                campaign_id=campaign_id,
                user=request.user if request.user.is_authenticated else None
            )
            return Response(result)

        except Exception as e:
            print(f"❌ Error in generate_campaign_analysis: {e}")
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ===========================================================
# REPORT TEMPLATE VIEWSET
# ===========================================================

class ReportTemplateViewSet(viewsets.ModelViewSet):
    """Manage report templates — campaign-scoped."""
    serializer_class = ReportTemplateSerializer
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        """FIX: filter by campaign_id when provided so templates don't leak across campaigns."""
        queryset = ReportTemplate.objects.filter(is_active=True)
        campaign_id = self.request.query_params.get('campaign_id')
        if campaign_id:
            queryset = queryset.filter(campaign_id=campaign_id)
        return queryset.order_by('-uploaded_at')

    def perform_create(self, serializer):
        serializer.save(
            uploaded_by=self.request.user if self.request.user.is_authenticated else None
        )

    def create(self, request, *args, **kwargs):
        """Handle template upload with sheet extraction."""
        print("=" * 50)
        print("📤 TEMPLATE UPLOAD REQUEST RECEIVED")
        print(f"Data keys: {list(request.data.keys())}")
        print(f"Campaign ID: {request.data.get('campaign_id') or request.data.get('campaign')}")

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()

        print(f"✅ Template ID={instance.id}, Campaign={instance.campaign}")

        # Extract sheet names after file is saved
        try:
            import time
            time.sleep(0.5)
            file_path = instance.template_file.path
            if os.path.exists(file_path):
                excel_file = pd.ExcelFile(file_path)
                instance.sheet_names = excel_file.sheet_names
                instance.save(update_fields=['sheet_names'])
                print(f"📑 Sheets: {instance.sheet_names}")
        except Exception as e:
            print(f"⚠️ Could not extract sheet names: {e}")

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=True, methods=['post'], url_path='extract-sheets', url_name='extract-sheets')
    def extract_sheets(self, request, pk=None):
        """Manually trigger sheet name extraction."""
        template = self.get_object()
        try:
            if not template.template_file:
                return Response({'success': False, 'error': 'No template file found'},
                                status=status.HTTP_400_BAD_REQUEST)
            file_path = template.template_file.path
            if not os.path.exists(file_path):
                return Response(
                    {'success': False, 'error': f'File not found at: {file_path}'},
                    status=status.HTTP_404_NOT_FOUND
                )
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            template.sheet_names = sheet_names
            template.save(update_fields=['sheet_names'])
            return Response({
                'success': True,
                'message': f'Extracted {len(sheet_names)} sheets',
                'sheets': sheet_names
            })
        except Exception as e:
            traceback.print_exc()
            return Response({'success': False, 'error': str(e)},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def configure_mapping(self, request, pk=None):
        """Configure sheet-to-data mappings."""
        template = self.get_object()
        sheet_mappings = request.data.get('sheet_mappings', {})

        if not template.sheet_names:
            return Response(
                {'error': 'Template has no sheet names. Extract sheets first.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        for sheet_name in sheet_mappings.values():
            if sheet_name and sheet_name not in template.sheet_names:
                return Response(
                    {'error': f"Sheet '{sheet_name}' not found. "
                              f"Available: {template.sheet_names}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        template.sheet_mappings = sheet_mappings
        template.save()
        return Response({
            'success': True,
            'message': 'Mapping configured successfully',
            'sheet_mappings': template.sheet_mappings
        })

    @action(detail=True, methods=['get'])
    def sheets(self, request, pk=None):
        """Return sheet names for a template."""
        template = self.get_object()
        return Response({
            'template_name': template.name,
            'sheets': template.sheet_names,
            'has_sheets': bool(template.sheet_names)
        })


@api_view(['POST'])
@permission_classes([AllowAny])
def upload_report_template(request):
    """Standalone endpoint to upload a report template."""
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=400)

    file = request.FILES['file']
    name = request.data.get('name', file.name)
    description = request.data.get('description', '')

    # FIX: resolve campaign
    campaign_id = request.data.get('campaign_id') or request.data.get('campaign')
    campaign = None
    if campaign_id:
        try:
            campaign = Campaign.objects.get(id=campaign_id, is_active=True)
        except Campaign.DoesNotExist:
            pass

    try:
        template = ReportTemplate.objects.create(
            name=name,
            description=description,
            template_file=file,
            campaign=campaign,
            uploaded_by=request.user if request.user.is_authenticated else None
        )

        try:
            import time
            time.sleep(0.5)
            file_path = template.template_file.path
            if os.path.exists(file_path):
                excel_file = pd.ExcelFile(file_path)
                sheet_names = excel_file.sheet_names
                template.sheet_names = sheet_names
                template.save(update_fields=['sheet_names'])
        except Exception as e:
            print(f"⚠️ Could not extract sheets: {e}")

        return Response({
            'success': True,
            'message': 'Template uploaded successfully',
            'template': {
                'id': template.id,
                'name': template.name,
                'campaign': campaign.display_name if campaign else None,
                'sheets': template.sheet_names or []
            }
        })
    except Exception as e:
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


# ===========================================================
# TEMPLATE-BASED REPORT GENERATOR
# ===========================================================

class TemplateBasedReportGenerator:
    """Generate reports by populating an uploaded Excel template with Pivot data."""

    @staticmethod
    def generate_analysis_report(template_id, data_mappings, campaign_id=None, user=None):
        """
        Generate campaign analysis by recreating the template structure,
        preserving ALL formulas, and populating matching rows with Pivot data.

        This restores the ORIGINAL working approach (which produced correct
        values), with ONE addition: campaign_id scoping, so the Pivot data
        used always belongs to the correct campaign — no more mixing data
        across campaigns.

        data_mappings: {
            'campaign': 'Sheet Name'  # The campaign sheet to populate
        }

        How it works:
        1. Takes the uploaded template and recreates it COMPLETELY
        2. PRESERVES ALL FORMULAS exactly as they are
        3. CLEARS ALL VALUES (except descriptions for matching)
        4. Takes ALL rows from the Pivot sheet data — SCOPED TO campaign_id
        5. For each row in the campaign sheet, looks for matching description
           in Pivot (column 1 = description)
        6. If match found, populates the ENTIRE row with Pivot data
           (so if Pivot has multiple columns, they all get copied across)
        7. If no match found, leaves the row with only the description
           (formulas preserved)
        8. ALL formulas remain in place, nothing is ever deleted
        """
        try:
            # Get template
            template = ReportTemplate.objects.get(id=template_id, is_active=True)

            print(f"📊 Generating campaign analysis using template: {template.name}")
            print(f"📑 Template sheets: {template.sheet_names}")
            print(f"📋 Data mappings: {data_mappings}")
            print(f"🎯 Campaign ID: {campaign_id}")

            # Get the campaign sheet name from mappings
            campaign_sheet_name = data_mappings.get('campaign')
            if not campaign_sheet_name:
                raise Exception("No campaign sheet specified in mappings")

            if not campaign_id:
                raise Exception(
                    "campaign_id is required so the correct campaign's Pivot "
                    "data is used (prevents mixing data across campaigns)."
                )

            # ── FIX: scope the campaign report lookup to THIS campaign ──
            latest_campaign_report = GeneratedReport.objects.filter(
                report_type='campaign_analysis',
                campaign_id=campaign_id
            ).order_by('-generated_at').first()

            if not latest_campaign_report:
                raise Exception(
                    f"No campaign report found for campaign ID {campaign_id}. "
                    "Please generate a campaign report for this campaign first."
                )

            print(f"📊 Using campaign report: {latest_campaign_report.id} "
                  f"(campaign_id={latest_campaign_report.campaign_id})")
            print(f"📁 Report file: {latest_campaign_report.file.path}")

            # STEP 1: Load the campaign report workbook and get its Pivot sheet
            campaign_wb = load_workbook(latest_campaign_report.file.path)

            if 'Pivot' not in campaign_wb.sheetnames:
                raise Exception("Pivot sheet not found in campaign report")

            pivot_sheet = campaign_wb['Pivot']
            print(f"📑 Pivot sheet has {pivot_sheet.max_row} rows and "
                  f"{pivot_sheet.max_column} columns")

            # STEP 2: Extract ALL Pivot data with descriptions as the key
            pivot_data_by_description = {}
            headers = []

            for i, row in enumerate(pivot_sheet.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(cell) if cell else f"Col_{i+1}" for i, cell in enumerate(row)]
                    print(f"📋 Pivot headers: {headers}")
                else:
                    if len(row) > 0 and row[0]:
                        description_key = str(row[0]).strip().lower()
                        pivot_data_by_description[description_key] = list(row)

            print(f"📊 Extracted {len(pivot_data_by_description)} rows of Pivot data "
                  f"with descriptions")
            if len(pivot_data_by_description) > 0:
                sample_key = list(pivot_data_by_description.keys())[0]
                print(f"📋 Sample - Description: '{sample_key}', "
                      f"Data: {pivot_data_by_description[sample_key]}")

            # ── STEP 2b: Inject SUMMARY METRICS as extra lookup entries ─────
            # The Pivot sheet only has disposition descriptions (Callback,
            # Not Interested, etc.) — it does NOT have rows for "Total Leads",
            # "Successful Contacts", "Conversion" etc. Those summary numbers
            # live in the Campaign Analysis sheet of the same campaign report,
            # at FIXED row numbers (this is exactly how generate_campaign()
            # writes them, so reading the same fixed rows is reliable).
            #
            #   row 4 (col B): Total Leads
            #   row 5 (col B): Unworked Leads (New Leads)
            #   row 6 (col B): Successful Contacts
            #   row 7 (col B): True Contacts
            #   row 8 (col B): Conversion (decimal, e.g. 0.0159)
            #   row 9 (col B): True Sales (post QA)
            #
            # We add these into pivot_data_by_description using the SAME
            # 2-column shape as Pivot rows: [description, value]. This lets
            # them flow through the exact same matching/writing logic below
            # as every disposition row — no separate code path needed.

            if 'Campaign Analysis' in campaign_wb.sheetnames:
                # Open a data_only copy so formula results (not formula text)
                # are what we read.
                value_wb = load_workbook(latest_campaign_report.file.path, data_only=True)
                analysis_value_ws = value_wb['Campaign Analysis']

                def read_summary_cell(row_num):
                    v = analysis_value_ws.cell(row=row_num, column=2).value
                    return v if isinstance(v, (int, float)) else 0

                summary_total_leads   = read_summary_cell(4)
                summary_unworked      = read_summary_cell(5)
                summary_successful    = read_summary_cell(6)
                summary_true_contacts = read_summary_cell(7)
                summary_conversion    = read_summary_cell(8)   # decimal
                summary_true_sales    = read_summary_cell(9)

                summary_entries = {
                    'total leads':                      summary_total_leads,
                    'unworked leads (new leads)':       summary_unworked,
                    'succesful contacts':               summary_successful,  # template typo (1 's')
                    'successful contacts':              summary_successful,
                    'true contacts':                    summary_true_contacts,
                    'conversion':                        summary_conversion,
                    'true sales (post qa)':             summary_true_sales,
                }

                for label, val in summary_entries.items():
                    # Use the SAME 2-column row shape as Pivot: [description, value]
                    pivot_data_by_description[label] = [label, val]

                print(f"📊 Injected {len(summary_entries)} summary metric entries:")
                for k, v in summary_entries.items():
                    print(f"   '{k}' = {v}")

                # ── Calculated ratio rows (Contactability, Lead to Sale, etc.) ──
                if summary_total_leads > 0:
                    contactability         = summary_successful / summary_total_leads
                    lead_to_sale_conv      = summary_true_sales / summary_total_leads
                    pivot_data_by_description['contactability'] = ['contactability', contactability]
                    pivot_data_by_description['lead to sale conversion'] = [
                        'lead to sale conversion', lead_to_sale_conv
                    ]
                if summary_successful > 0:
                    conv_to_true_sale = summary_true_sales / summary_successful
                    pivot_data_by_description['conversion to true sale'] = [
                        'conversion to true sale', conv_to_true_sale
                    ]
                if summary_true_contacts > 0:
                    true_contacts_to_sales = summary_true_sales / summary_true_contacts
                    pivot_data_by_description['true contacts to sales'] = [
                        'true contacts to sales', true_contacts_to_sales
                    ]

                print(f"📊 pivot_data_by_description now has "
                      f"{len(pivot_data_by_description)} total entries "
                      f"(disposition rows + summary metrics + calculated ratios)")
            else:
                print("⚠️  'Campaign Analysis' sheet not found in campaign report — "
                      "summary metrics (Total Leads, Successful Contacts, etc.) "
                      "will be left blank.")

            # STEP 3: Load the template workbook (this is our source template)
            template_path = template.template_file.path
            source_wb = load_workbook(template_path)

            # STEP 4: Create a NEW workbook with the COMPLETE template structure
            new_wb = openpyxl.Workbook()
            default_sheet = new_wb.active
            new_wb.remove(default_sheet)

            print(f"📝 Recreating template structure with {len(source_wb.sheetnames)} "
                  f"sheets (preserving formulas)...")

            for sheet_name in source_wb.sheetnames:
                source_sheet = source_wb[sheet_name]
                new_sheet = new_wb.create_sheet(title=sheet_name)

                for row in source_sheet.iter_rows():
                    for cell in row:
                        new_cell = new_sheet.cell(row=cell.row, column=cell.column)

                        if cell.has_style:
                            try:
                                new_cell.font = cell.font.copy()
                                new_cell.border = cell.border.copy()
                                new_cell.fill = cell.fill.copy()
                                new_cell.number_format = cell.number_format
                                new_cell.alignment = cell.alignment.copy()
                            except Exception:
                                pass

                        # Preserve formulas
                        if cell.data_type == 'f' and cell.value and str(cell.value).startswith('='):
                            new_cell.value = cell.value
                        else:
                            # For the campaign sheet, preserve description column (col 1) values
                            if sheet_name == campaign_sheet_name and cell.column == 1:
                                new_cell.value = cell.value
                            else:
                                new_cell.value = None  # clear all other values

                for merged_range in source_sheet.merged_cells.ranges:
                    new_sheet.merge_cells(str(merged_range))

                for col_idx, column in enumerate(source_sheet.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    if col_letter in source_sheet.column_dimensions:
                        new_sheet.column_dimensions[col_letter].width = (
                            source_sheet.column_dimensions[col_letter].width
                        )

                for row_idx in range(1, source_sheet.max_row + 1):
                    if row_idx in source_sheet.row_dimensions:
                        new_sheet.row_dimensions[row_idx].height = (
                            source_sheet.row_dimensions[row_idx].height
                        )

                print(f"  ✅ Recreated sheet: {sheet_name} - {source_sheet.max_row} rows "
                      f"(formulas preserved)")

            print(f"✅ Template structure successfully recreated with "
                  f"{len(source_wb.sheetnames)} sheets (all formulas preserved)")

            # STEP 5: Get the campaign sheet from the NEW workbook
            if campaign_sheet_name not in new_wb.sheetnames:
                raise Exception(
                    f"Campaign sheet '{campaign_sheet_name}' not found in recreated template"
                )

            campaign_sheet = new_wb[campaign_sheet_name]

            # STEP 6: Description column is column 1
            description_column = 1
            total_rows = campaign_sheet.max_row

            print(f"📝 Processing campaign sheet '{campaign_sheet_name}' with "
                  f"{total_rows} rows...")

            # STEP 7: Populate ONLY matching rows with data from Pivot
            rows_populated = 0
            rows_without_match = 0

            for row_idx in range(1, total_rows + 1):
                desc_cell = campaign_sheet.cell(row=row_idx, column=description_column)
                description = str(desc_cell.value).strip().lower() if desc_cell.value else ""

                if description and description in pivot_data_by_description:
                    pivot_row_data = pivot_data_by_description[description]
                    rows_populated += 1

                    print(f"  ✅ Matching row {row_idx}: '{description}' - "
                          f"populating with Pivot data")

                    for col_idx, value in enumerate(pivot_row_data, start=1):
                        if col_idx <= campaign_sheet.max_column:
                            cell = campaign_sheet.cell(row=row_idx, column=col_idx)

                            # Skip formula cells — never overwrite formulas
                            if cell.data_type == 'f':
                                continue

                            try:
                                if value is None:
                                    cell.value = None
                                elif isinstance(value, float) and np.isnan(value):
                                    cell.value = None
                                elif isinstance(value, (np.integer, int)):
                                    cell.value = int(value)
                                elif isinstance(value, (np.floating, float)):
                                    cell.value = float(value)
                                elif isinstance(value, (pd.Timestamp, datetime)):
                                    cell.value = value
                                else:
                                    cell.value = str(value)
                            except Exception as e:
                                print(f"⚠️ Error writing to cell ({row_idx},{col_idx}): {e}")
                else:
                    rows_without_match += 1
                    if description:
                        print(f"  ⚠️ No match for row {row_idx}: '{description}' - "
                              f"keeping formulas only")

            print(f"✅ Campaign sheet '{campaign_sheet_name}' results:")
            print(f"   - Total rows: {total_rows}")
            print(f"   - Rows populated with Pivot data: {rows_populated}")
            print(f"   - Rows with formulas only (no match): {rows_without_match}")

            # STEP 8: Add the original Pivot sheet as a separate sheet for reference
            pivot_data_sheet = new_wb.create_sheet(title="Pivot_Data")

            for row in pivot_sheet.iter_rows():
                for cell in row:
                    new_cell = pivot_data_sheet.cell(row=cell.row, column=cell.column)
                    new_cell.value = cell.value
                    if cell.has_style:
                        try:
                            new_cell.font = cell.font.copy()
                            new_cell.border = cell.border.copy()
                            new_cell.fill = cell.fill.copy()
                            new_cell.number_format = cell.number_format
                            new_cell.alignment = cell.alignment.copy()
                        except Exception:
                            pass

            for merged_range in pivot_sheet.merged_cells.ranges:
                pivot_data_sheet.merge_cells(str(merged_range))

            for col_idx, column in enumerate(pivot_sheet.columns, 1):
                col_letter = get_column_letter(col_idx)
                if col_letter in pivot_sheet.column_dimensions:
                    pivot_data_sheet.column_dimensions[col_letter].width = (
                        pivot_sheet.column_dimensions[col_letter].width
                    )

            print(f"✅ Added 'Pivot_Data' sheet with {pivot_sheet.max_row-1} rows "
                  f"of reference data")

            # STEP 8b: Also copy "Processed Data" and "Campaign Analysis" sheets
            # from the campaign report, so the final download has the full
            # 4-sheet set: Processed Data, Pivot_Data, Campaign Analysis,
            # and the populated template sheet.
            for extra_sheet_name in ['Processed Data', 'Campaign Analysis']:
                if extra_sheet_name in campaign_wb.sheetnames and extra_sheet_name not in new_wb.sheetnames:
                    src_extra = campaign_wb[extra_sheet_name]
                    dst_extra = new_wb.create_sheet(title=extra_sheet_name)

                    for row in src_extra.iter_rows():
                        for cell in row:
                            new_cell = dst_extra.cell(row=cell.row, column=cell.column)
                            new_cell.value = cell.value
                            if cell.has_style:
                                try:
                                    new_cell.font = cell.font.copy()
                                    new_cell.border = cell.border.copy()
                                    new_cell.fill = cell.fill.copy()
                                    new_cell.number_format = cell.number_format
                                    new_cell.alignment = cell.alignment.copy()
                                except Exception:
                                    pass

                    for merged_range in src_extra.merged_cells.ranges:
                        dst_extra.merge_cells(str(merged_range))

                    for col_idx, column in enumerate(src_extra.columns, 1):
                        col_letter = get_column_letter(col_idx)
                        if col_letter in src_extra.column_dimensions:
                            dst_extra.column_dimensions[col_letter].width = (
                                src_extra.column_dimensions[col_letter].width
                            )

                    print(f"✅ Added '{extra_sheet_name}' sheet "
                          f"({src_extra.max_row} rows) for reference")

            # STEP 9: Save the new workbook
            output = BytesIO()
            new_wb.save(output)
            output.seek(0)

            reports_dir = os.path.join(settings.MEDIA_ROOT, 'campaign_analysis')
            os.makedirs(reports_dir, exist_ok=True)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"campaign_analysis_{campaign_sheet_name}_{timestamp}.xlsx"
            filepath = os.path.join(reports_dir, filename)

            with open(filepath, 'wb') as f:
                f.write(output.getvalue())

            campaign_obj = None
            try:
                campaign_obj = Campaign.objects.get(id=campaign_id)
            except Exception:
                pass

            # Create report record — FIX: campaign attached for scoping
            report = GeneratedReport.objects.create(
                user=user,
                campaign=campaign_obj,
                report_type='analysis',
                file=f"campaign_analysis/{filename}",
                parameters={
                    'template_id': template.id,
                    'template_name': template.name,
                    'campaign_id': campaign_id,
                    'report_name': f"Analysis - {campaign_sheet_name} - "
                                    f"{datetime.now().strftime('%Y-%m-%d')}",
                    'campaign_name': campaign_sheet_name,
                    'source_report_id': latest_campaign_report.id,
                    'pivot_data_sheet': 'Pivot_Data',
                    'total_template_rows': total_rows,
                    'rows_populated': rows_populated,
                    'rows_with_formulas_only': rows_without_match,
                    'template_sheets': source_wb.sheetnames,
                    'sheets_included': new_wb.sheetnames,
                    'all_formulas_preserved': True,
                    'values_cleared': True
                }
            )

            return {
                'success': True,
                'data': {
                    'report_id': report.id,
                    'download_url': f'/api/reports/{report.id}/download/',
                    'message': (
                        f'Campaign analysis generated for "{campaign_sheet_name}" - '
                        f'Populated {rows_populated} rows with Pivot data, '
                        f'{rows_without_match} rows with formulas only'
                    ),
                    'campaign_name': campaign_sheet_name,
                    'rows_populated': rows_populated,
                    'rows_formulas_only': rows_without_match,
                    'total_rows': total_rows,
                    'pivot_sheet': 'Pivot_Data',
                    'sheets': new_wb.sheetnames,
                }
            }

        except ReportTemplate.DoesNotExist:
            raise Exception(f"Template with ID {template_id} not found")
        except Exception as e:
            print(f"❌ Error generating campaign analysis: {str(e)}")
            traceback.print_exc()
            raise


    @staticmethod
    def _clear_data_preserve_structure(sheet, start_row=2):
        """
        Placeholder — we no longer pre-clear cells before writing.
        The write step handles preservation by checking each cell individually.
        Kept so any external callers don't break.
        """
        print(f"📝 Preserving all template content — only data cells will be updated")

    @staticmethod
    def _write_data_preserve_structure(sheet, df, start_row=2, start_col=1):
        """
        Write a DataFrame to a sheet while preserving the EXACT template structure.

        Rules:
        - NEVER removes or adds rows — template structure is sacred.
        - Never writes beyond the template's existing row count.
        - Skips formula cells (preserves them unchanged).
        - Skips non-master cells of merged ranges.
        - For empty (NaN/None) values, only clears the cell if it was already empty;
          existing template text is left alone.
        """
        if df.empty:
            print("⚠️ No data to write — template remains unchanged")
            return

        merged_ranges = list(sheet.merged_cells.ranges)
        merged_top_left_cells = {
            mr.start_cell.coordinate: mr for mr in merged_ranges
        }

        max_template_rows = sheet.max_row
        data_rows = len(df)

        print(f"📝 Template rows: {max_template_rows}  |  Data rows: {data_rows}  |  Start: {start_row}")

        for i in range(data_rows):
            row_idx = start_row + i

            if row_idx > max_template_rows:
                print(f"⚠️ Stopped at row {row_idx} — template only has {max_template_rows} rows. "
                      f"{data_rows - i} extra data rows NOT written.")
                break

            row_data = df.iloc[i]

            for col_offset, value in enumerate(row_data):
                col_idx = start_col + col_offset
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_coord = cell.coordinate

                # Never overwrite formula cells
                if cell.data_type == 'f':
                    continue

                # Determine if this cell is inside a merged range (but not the master)
                is_non_master_merged = False
                if cell_coord not in merged_top_left_cells:
                    for mr in merged_ranges:
                        if cell_coord in mr and cell_coord != mr.start_cell.coordinate:
                            is_non_master_merged = True
                            break

                if is_non_master_merged:
                    continue  # can't write to non-master merged cells

                try:
                    is_empty_value = (
                        value is None
                        or (isinstance(value, float) and np.isnan(value))
                        or (isinstance(value, str) and value.strip() == '')
                    )

                    if is_empty_value:
                        # Only clear if the cell has no existing template content
                        if cell.value is None:
                            cell.value = None
                        # else: preserve existing template content
                    else:
                        if isinstance(value, (np.integer, int)):
                            cell.value = int(value)
                        elif isinstance(value, (np.floating, float)):
                            cell.value = float(value)
                        elif isinstance(value, (pd.Timestamp, datetime)):
                            cell.value = value
                        else:
                            cell.value = str(value)

                except Exception as e:
                    print(f"⚠️ Error writing to cell {cell_coord}: {e}")

    @staticmethod
    def get_available_sheets(template_id):
        try:
            template = ReportTemplate.objects.get(id=template_id, is_active=True)
            return template.sheet_names
        except ReportTemplate.DoesNotExist:
            return []


# ===========================================================
# DASHBOARD STATS VIEW
# ===========================================================

class DashboardStatsView(generics.GenericAPIView):
    """Get dashboard statistics, optionally scoped to a campaign."""
    permission_classes = [AllowAny]

    def get(self, request):
        user = request.user
        campaign_id = request.query_params.get('campaign_id')

        total_outcomes = OutcomeDescription.objects.count()

        base_files_qs   = CallDataFile.objects.all()
        base_reports_qs = GeneratedReport.objects.all()

        if campaign_id:
            base_files_qs   = base_files_qs.filter(campaign_id=campaign_id)
            base_reports_qs = base_reports_qs.filter(campaign_id=campaign_id)
        elif user.is_authenticated:
            base_files_qs   = base_files_qs.filter(user=user)
            base_reports_qs = base_reports_qs.filter(user=user)

        total_files     = base_files_qs.count()
        processed_files = base_files_qs.filter(status='processed').count()
        total_reports   = base_reports_qs.count()

        total_records = sum(
            f.total_records for f in base_files_qs.filter(status='processed')
        )

        stats = {
            'overview': {
                'total_outcomes':  total_outcomes,
                'total_files':     total_files,
                'processed_files': processed_files,
                'total_reports':   total_reports,
                'total_records':   total_records,
            },
            'recent_reports': GeneratedReportSerializer(
                base_reports_qs[:5], many=True
            ).data,
            'recent_files': CallDataFileSerializer(
                base_files_qs[:5], many=True
            ).data,
            'file_status': {
                'uploaded':   base_files_qs.filter(status='uploaded').count(),
                'processing': base_files_qs.filter(status='processing').count(),
                'processed':  processed_files,
                'failed':     base_files_qs.filter(status='failed').count(),
            }
        }
        return Response(stats)


# ===========================================================
# CAMPAIGN VIEWSET
# ===========================================================

class CampaignViewSet(viewsets.ModelViewSet):
    """Manage campaigns."""
    queryset = Campaign.objects.filter(is_active=True).order_by('name')
    serializer_class = CampaignSerializer
    permission_classes = [AllowAny]

    def perform_create(self, serializer):
        serializer.save(
            created_by=self.request.user if self.request.user.is_authenticated else None
        )

    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        """Campaign-specific statistics."""
        campaign = self.get_object()
        return Response({
            'campaign': {
                'id':         campaign.id,
                'name':       campaign.display_name,
                'sheet_name': campaign.sheet_name,
            },
            'data_files': {
                'total':      campaign.data_files.count(),
                'processed':  campaign.data_files.filter(status='processed').count(),
                'uploaded':   campaign.data_files.filter(status='uploaded').count(),
                'processing': campaign.data_files.filter(status='processing').count(),
                'failed':     campaign.data_files.filter(status='failed').count(),
            },
            'reports': {
                'total':             campaign.reports.count(),
                'campaign_analysis': campaign.reports.filter(
                    report_type='campaign_analysis'
                ).count(),
                'template_analysis': campaign.reports.filter(
                    report_type='analysis'
                ).count(),
            },
            'templates':     campaign.templates.count(),
            'total_records': sum(
                f.total_records
                for f in campaign.data_files.filter(status='processed')
            ),
        })

    @action(detail=True, methods=['get'])
    def recent_activity(self, request, pk=None):
        """Recent files and reports for this campaign."""
        campaign = self.get_object()
        recent_files   = CallDataFile.objects.filter(
            campaign=campaign
        ).order_by('-uploaded_at')[:5]
        recent_reports = GeneratedReport.objects.filter(
            campaign=campaign
        ).order_by('-generated_at')[:5]
        return Response({
            'recent_files':   CallDataFileSerializer(recent_files, many=True).data,
            'recent_reports': GeneratedReportSerializer(recent_reports, many=True).data,
        })

    @action(detail=True, methods=['post'])
    def sync_from_database(self, request, pk=None):
        """Pull this campaign's call data from the external source database."""
        from .external_source import sync_campaign_from_database, ExternalSourceError

        campaign = self.get_object()
        user = request.user if request.user.is_authenticated else None
        try:
            instance = sync_campaign_from_database(campaign, user=user)
        except ExternalSourceError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f'Database sync failed: {e}'}, status=status.HTTP_502_BAD_GATEWAY)

        return Response(CallDataFileSerializer(instance).data)


# ===========================================================
# UTILITY VIEWS
# ===========================================================

@api_view(['POST'])
@permission_classes([AllowAny])
def test_upload(request):
    """Test endpoint for file upload."""
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=400)
    file_obj = request.FILES['file']
    try:
        df = pd.read_excel(file_obj)
        return Response({
            'success':   True,
            'filename':  file_obj.name,
            'columns':   list(df.columns),
            'row_count': len(df)
        })
    except Exception as e:
        return Response({'error': str(e)}, status=400)


@api_view(['GET'])
@permission_classes([AllowAny])
def setup_test_user(request):
    """Create a test user for quick setup."""
    if not User.objects.filter(username='test').exists():
        user = User.objects.create_user(
            username='test', email='test@example.com', password='test123'
        )
        user.is_active = True
        user.save()
        return Response({
            'message': 'Test user created successfully!',
            'credentials': {'username': 'test', 'password': 'test123'}
        })
    return Response({
        'message': 'Test user already exists',
        'credentials': {'username': 'test', 'password': 'test123'}
    })